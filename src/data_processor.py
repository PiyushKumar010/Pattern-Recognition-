import pandas as pd
from openpyxl import load_workbook
import hashlib
from datetime import datetime

# PostgreSQL connection utilities
from src.utils.db import get_db_cursor, get_db_connection, execute_query

def fetch_data_from_db(query, params=None):
    """
    Fetch data from PostgreSQL database with proper connection management.
    
    Args:
        query (str): SQL query to execute (use %s for parameters)
        params (tuple/dict): Query parameters for safe parameterized queries
    
    Returns:
        list of tuples: Query results
        
    Example:
        # Simple query
        results = fetch_data_from_db("SELECT * FROM users WHERE active = %s", (True,))
        
        # Query with multiple parameters
        results = fetch_data_from_db(
            "SELECT * FROM orders WHERE user_id = %s AND date > %s",
            (user_id, start_date)
        )
    """
    try:
        return execute_query(query, params, fetch=True)
    except Exception as e:
        print(f"Database query error: {e}")
        return None


def fetch_dataframe_from_db(query, params=None):
    """
    Fetch data from PostgreSQL and return as pandas DataFrame.
    
    Args:
        query (str): SQL query to execute
        params (tuple/dict): Query parameters
    
    Returns:
        pd.DataFrame: Query results as DataFrame
    """
    try:
        with get_db_connection() as conn:
            df = pd.read_sql_query(query, conn, params=params)
        return df
    except Exception as e:
        print(f"Database DataFrame query error: {e}")
        return None


def save_dataframe_to_db(df, table_name, if_exists='replace', index=False):
    """
    Save pandas DataFrame to PostgreSQL table using the FASTEST method.
    Uses PostgreSQL COPY command for maximum performance (10-50x faster than INSERT).
    
    Args:
        df (pd.DataFrame): DataFrame to save
        table_name (str): Name of the database table
        if_exists (str): 'fail', 'replace', or 'append' (default: 'replace')
        index (bool): Write DataFrame index as a column
    
    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        from sqlalchemy import create_engine, text
        from io import StringIO
        import sys
        import os
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
        from config import POSTGRESQL_CONFIG
        
        print(f"Starting optimized upload for {len(df)} rows...")
        
        # Create SQLAlchemy engine
        connection_string = f"postgresql://{POSTGRESQL_CONFIG['user']}:{POSTGRESQL_CONFIG['password']}@{POSTGRESQL_CONFIG['host']}:{POSTGRESQL_CONFIG['port']}/{POSTGRESQL_CONFIG['database']}"
        engine = create_engine(connection_string)
        
        # Handle if_exists parameter
        with engine.connect() as connection:
            if if_exists == 'replace':
                # Drop table if exists
                connection.execute(text(f'DROP TABLE IF EXISTS "{table_name}"'))
                connection.commit()
                print("Dropped existing table (if any)")
            elif if_exists == 'fail':
                # Check if table exists
                result = connection.execute(text(
                    "SELECT EXISTS (SELECT FROM pg_tables WHERE tablename = :table_name)"
                ), {"table_name": table_name})
                if result.fetchone()[0]:
                    raise ValueError(f"Table '{table_name}' already exists")
        
        # Create table structure using to_sql with 0 rows (fast)
        print("Creating table structure...")
        df.head(0).to_sql(table_name, engine, if_exists='append', index=index)
        
        # Now use PostgreSQL COPY for bulk insert (FAST!)
        print(f"Bulk loading {len(df)} rows using PostgreSQL COPY command...")
        
        # Get raw psycopg2 connection
        from src.utils.db import get_postgres_connection, return_connection
        conn = get_postgres_connection()
        
        try:
            # Convert DataFrame to CSV in memory
            csv_buffer = StringIO()
            df.to_csv(csv_buffer, index=index, header=False, sep='\t', na_rep='\\N')
            csv_buffer.seek(0)
            
            # Use COPY command (PostgreSQL's fastest bulk loader)
            cursor = conn.cursor()
            
            # Build column list
            columns = ', '.join([f'"{col}"' for col in df.columns])
            
            copy_sql = f'COPY "{table_name}" ({columns}) FROM STDIN WITH (FORMAT CSV, DELIMITER E\'\\t\', NULL \'\\N\')'
            cursor.copy_expert(copy_sql, csv_buffer)
            
            conn.commit()
            cursor.close()
            
            message = f"✅ Successfully uploaded {len(df):,} rows to table '{table_name}' using optimized COPY method"
            print(message)
            
            return True, message
            
        finally:
            return_connection(conn)
        
    except Exception as e:
        error_msg = f"Error saving DataFrame to database: {e}"
        print(error_msg)
        
        # Fallback to slower method if COPY fails
        print("Falling back to standard method...")
        try:
            from sqlalchemy import create_engine
            connection_string = f"postgresql://{POSTGRESQL_CONFIG['user']}:{POSTGRESQL_CONFIG['password']}@{POSTGRESQL_CONFIG['host']}:{POSTGRESQL_CONFIG['port']}/{POSTGRESQL_CONFIG['database']}"
            engine = create_engine(connection_string)
            
            df.to_sql(table_name, engine, if_exists=if_exists, index=index, method='multi', chunksize=5000)
            message = f"Successfully saved {len(df)} rows using fallback method"
            print(message)
            return True, message
        except Exception as fallback_error:
            return False, f"Both methods failed: {str(fallback_error)}"


def get_table_name_from_file(filename):
    """
    Generate a unique table name from uploaded file name.
    Uses hash to ensure uniqueness and avoid SQL injection.
    
    Args:
        filename (str): Original filename
    
    Returns:
        str: Safe table name
    """
    # Create hash of filename with timestamp for uniqueness
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    hash_obj = hashlib.md5(f"{filename}_{timestamp}".encode())
    hash_str = hash_obj.hexdigest()[:8]
    
    # Clean filename to create base name
    base_name = filename.replace('.xlsx', '').replace('.xls', '').replace(' ', '_').replace('-', '_')
    # Remove special characters
    base_name = ''.join(c for c in base_name if c.isalnum() or c == '_')
    base_name = base_name[:20]  # Limit length
    
    return f"upload_{base_name}_{hash_str}".lower()


def get_column_info_from_table(table_name):
    """
    Get column information from a database table.
    
    Args:
        table_name (str): Name of the table
    
    Returns:
        dict: Column information with types
    """
    try:
        query = """
            SELECT column_name, data_type 
            FROM information_schema.columns 
            WHERE table_name = %s
            ORDER BY ordinal_position
        """
        results = execute_query(query, (table_name,), fetch=True)
        
        column_info = {}
        for col_name, data_type in results:
            column_info[col_name] = data_type
        
        return column_info
    except Exception as e:
        print(f"Error getting column info: {e}")
        return {}

def load_and_process_data(uploaded_file, save_to_db=True):
    """
    Load the uploaded Excel or CSV file, process it, and optionally save to database.
    Returns both the DataFrame and the table name if saved to database.
    
    Args:
        uploaded_file: Uploaded file object
        save_to_db (bool): Whether to save data to PostgreSQL database
    
    Returns:
        tuple: (df, table_name) if save_to_db=True, else (df, None)
    """
    try:
        # Detect file type
        uploaded_file.seek(0)
        filename = getattr(uploaded_file, 'name', 'uploaded_file')
        is_csv = filename.lower().endswith('.csv')
        
        if is_csv:
            # Handle CSV files
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file)
            date_column_names = []
            
            print(f"After reading CSV: {len(df)} rows")
            
            # Convert numeric columns
            for col in df.columns:
                sample_values = df[col].dropna()
                if len(sample_values) == 0:
                    continue
                
                # Try to convert to numeric if possible
                try:
                    numeric_converted = pd.to_numeric(df[col], errors='coerce')
                    if numeric_converted.notna().sum() > 0:
                        non_null_count = len(df[col].dropna())
                        if non_null_count > 0 and numeric_converted.notna().sum() / non_null_count > 0.8:
                            df[col] = numeric_converted
                except:
                    pass
            
        else:
            # Handle Excel files
            uploaded_file.seek(0)  # Reset file pointer
            wb = load_workbook(uploaded_file, data_only=True)
            ws = wb.active
        
            # Track which columns are date columns (by column index)
            date_column_indices = set()
            
            # Read data preserving original cell formats
            data = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=False)):
                row_data = []
                for col_idx, cell in enumerate(row):
                    # Get the formatted value as it appears in Excel
                    if cell.value is not None:
                        # If cell has a number format (date/time), use the displayed value
                        if hasattr(cell, 'number_format') and cell.number_format and cell.number_format != 'General':
                            # Try to get the formatted string value
                            try:
                                from openpyxl.styles.numbers import is_date_format
                                if is_date_format(cell.number_format):
                                    # Track this column as a date column
                                    date_column_indices.add(col_idx)
                                    
                                    # For date cells, format them according to their Excel format
                                    if isinstance(cell.value, (int, float)):
                                        # Excel serial date - convert to datetime then format
                                        from datetime import datetime, timedelta
                                        excel_date = datetime(1899, 12, 30) + timedelta(days=cell.value)
                                        # Format based on number_format
                                        if 'h' in cell.number_format.lower() or 'm' in cell.number_format.lower() or 's' in cell.number_format.lower():
                                            # Has time component
                                            row_data.append(excel_date.strftime('%d-%m-%Y %H:%M:%S'))
                                        else:
                                            # Date only
                                            row_data.append(excel_date.strftime('%d-%m-%Y'))
                                    else:
                                        # Already a datetime object
                                        if 'h' in cell.number_format.lower() or 'm' in cell.number_format.lower() or 's' in cell.number_format.lower():
                                            row_data.append(cell.value.strftime('%d-%m-%Y %H:%M:%S'))
                                        else:
                                            row_data.append(cell.value.strftime('%d-%m-%Y'))
                                else:
                                    row_data.append(cell.value)
                            except:
                                row_data.append(cell.value)
                        else:
                            row_data.append(cell.value)
                    else:
                        row_data.append(None)
                data.append(row_data)
            
            # Convert to DataFrame
            df = pd.DataFrame(data[1:], columns=data[0])  # First row is header
            
            # Store date column names as metadata
            date_column_names = [data[0][idx] for idx in date_column_indices if idx < len(data[0])]
        
        # Common processing for both CSV and Excel
        df.attrs['date_columns'] = date_column_names
        
        print(f"After reading file: {len(df)} rows")
        print(f"Detected date columns: {date_column_names}")
        
        empty_rows = df.isnull().all(axis=1).sum()
        print(f"Completely empty rows found: {empty_rows}")
        
        # Convert pure numeric columns to numeric types (only for Excel files)
        if not is_csv:
            for col in df.columns:
                # Skip if column is empty
                sample_values = df[col].dropna()
                if len(sample_values) == 0:
                    continue
                
                # Skip columns that look like dates (already preserved as strings)
                sample_str = str(sample_values.iloc[0]) if len(sample_values) > 0 else ""
                if '-' in sample_str or '/' in sample_str or ':' in sample_str:
                    continue
                
                # Try to convert to numeric if possible (for pure numeric columns)
                try:
                    numeric_converted = pd.to_numeric(df[col], errors='coerce')
                    # Only convert if at least some values are actually numeric
                    if numeric_converted.notna().sum() > 0:
                        # Check if conversion makes sense (not all NaN after conversion)
                        non_null_count = len(df[col].dropna())
                        if non_null_count > 0 and numeric_converted.notna().sum() / non_null_count > 0.8:
                            df[col] = numeric_converted
                except:
                    pass  # Keep as string if conversion fails
        
        # Save to database if requested
        table_name = None
        dataset_id = None
        if save_to_db:
            # Generate unique table name
            filename = getattr(uploaded_file, 'name', 'uploaded_file')
            table_name = get_table_name_from_file(filename)
            
            # Save DataFrame to database
            success, message = save_dataframe_to_db(df, table_name, if_exists='replace')
            if success:
                print(f"Data saved to database table: {table_name}")
                
                # Save dataset metadata to history system
                try:
                    from src.history_manager import save_dataset, generate_file_hash
                    
                    # Get file content for hash
                    uploaded_file.seek(0)
                    file_content = uploaded_file.read()
                    file_hash = generate_file_hash(file_content)
                    file_size = len(file_content)
                    uploaded_file.seek(0)  # Reset for potential reuse
                    
                    dataset_id = save_dataset(
                        original_filename=filename,
                        table_name=table_name,
                        file_hash=file_hash,
                        file_size_bytes=file_size,
                        rows_count=len(df),
                        columns_count=len(df.columns)
                    )
                    if dataset_id:
                        print(f"Dataset metadata saved with ID: {dataset_id}")
                except Exception as e:
                    print(f"Warning: Failed to save dataset metadata: {e}")
            else:
                print(f"Warning: Failed to save to database: {message}")
                table_name = None
        
        # Store dataset_id in df attributes for later use
        if dataset_id:
            df.attrs['dataset_id'] = dataset_id
        
        return df, table_name
        
    except Exception as e:
        raise Exception(f"Error loading file: {str(e)}")